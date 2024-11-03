import time
import openpyxl
from selenium import webdriver
from utils import perform_action, perform_assertion
from report_generator import generate_report
import sys
from config import Config

class TestExecutor:
    def __init__(self, driver_path, excel_files, environment):
        self.config = Config.get_config(environment)
        self.driver = webdriver.Chrome(driver_path)
        self.excel_files = excel_files
        self.base_url = self.config['base_url']
        self.timeout = self.config['timeout']
        self.results = []

    def run_tests(self):
        self.driver.get(self.base_url)
        for excel_file in self.excel_files:
            workbook = openpyxl.load_workbook(excel_file)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    step, keyword, element_xpath, assertion_type, value, assertion, assertion_action, timeout = row
                    result = self.execute_step(keyword, element_xpath, assertion_type, value, assertion, assertion_action, timeout)
                    self.results.append((step, result))

        generate_report(self.results)

    def execute_step(self, keyword, element_xpath, assertion_type, value, assertion, assertion_action, timeout):
        try:
            perform_action(self.driver, keyword, element_xpath, timeout)
            result = perform_assertion(self.driver, assertion_type, value, assertion, assertion_action)
            return result
        except Exception as e:
            return f"Failed: {str(e)}"
        finally:
            time.sleep(1)  # Delay for stability

    def close(self):
        self.driver.quit()

if __name__ == "__main__":

    if len(sys.argv) != 2:
        print("Usage: python test_executor.py [dev|stage]")
        sys.exit(1)
    environment = sys.argv[1]
    executor = TestExecutor('path/to/chromedriver', ['tests/regression_test.xlsx', 'tests/sanity_test.xlsx'], environment)
    executor.run_tests()
    executor.close()